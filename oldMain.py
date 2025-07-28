import configparser
import pandas as pd
import time
import os.path
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from fake_useragent import UserAgent
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import keyring
import pickle
import re
import undetected_chromedriver as uc
import random

#setup the driver
def setupDriver():
    options = Options()
    options.add_argument(f'user-agent={UserAgent().random}')
    options.add_argument("--start-maximized")
    #fill in this line if you want to use your own user data
    #options.add_argument("user-data-dir=")
    driver = uc.Chrome(options=options, executable_path='chromedriver.exe')
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.implicitly_wait(10)
    return driver

#logins the user to the website
def login(driver, config):
    driver.find_element(By.XPATH, config.get('loginPath', 'userName')).send_keys(keyring.get_credential(service_name='Onboard', username=None).username) 
    driver.find_element(By.XPATH, config.get('loginPath', 'password')).send_keys(keyring.get_credential(service_name='Onboard', username=None).password)
    driver.find_element(By.XPATH, config.get('loginPath', 'loginB')).click()
    haltStep(1)

#makes sure an element loads before clicking it
def waitToLoad(driver, byType, identifier):
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((byType, identifier)))
    except:
        TimeoutException

#randomly stops the script between 1-3 seconds (to avoid detection)
def haltStep(x):
    #int = random.randint(1, 3)
    time.sleep(x)

def is_element_present(driver, by_locator, value):
    """
    Checks if an element is present in the DOM without raising an exception.
    Returns the element if found, None otherwise.
    """
    try:
        return driver.find_element(by_locator, value)
    except NoSuchElementException:
        return None

def writeOut(row_manual_review):
    if row_manual_review:
        df_manual_review = pd.DataFrame(row_manual_review)
        output_excel_path = r'C:\Users\dmartinez\Documents\ope\tenants_for_manual_review.xlsx' # Define your output path

        if os.path.exists(output_excel_path):
            try:
                with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Check if the sheet 'Manual Review' exists in the writer's book
                    # If it exists, find the next available row. Otherwise, start from row 0.
                    start_row = 0
                    if 'Manual Review' in writer.book.sheetnames:
                        # Get the sheet object to find its max_row
                        sheet = writer.book['Manual Review']
                        start_row = sheet.max_row
                        df_manual_review.to_excel(writer, sheet_name='Manual Review', index=False, header=False, startrow=start_row)
                    else:
                        # If the sheet doesn't exist, create it with headers at the top
                        df_manual_review.to_excel(writer, sheet_name='Manual Review', index=False, header=True, startrow=0)

                print(f"\nSuccessfully appended {len(row_manual_review)} rows requiring manual intervention to '{output_excel_path}'")
            except Exception as e:
                print(f"Error appending data to Excel: {e}")
                # Fallback to overwrite if append fails, or handle as per your needs
                try:
                    df_manual_review.to_excel(output_excel_path, index=False)
                    print(f"Fallback: Overwrote existing file due to append error. Check '{output_excel_path}'")
                except Exception as e_fallback:
                    print(f"Error overwriting file as fallback: {e_fallback}")
        else:
            # If the file doesn't exist, create it normally (this is equivalent to mode='w')
            try:
                df_manual_review.to_excel(output_excel_path, sheet_name='Manual Review', index=False, header=True)
                print(f"\nSuccessfully created '{output_excel_path}' with {len(row_manual_review)} rows requiring manual intervention.")
            except Exception as e:
                print(f"Error creating new Excel file: {e}")
    else:
        print("\nNo rows required manual intervention in this run.")

def main():
    idRegex = r'id="(.*?)"'
    #setup the config parser (where the html identifiers are stored)
    config = configparser.ConfigParser()
    config.read('venv\settings.ini')

    reactorLogin = config.get('urlPath', 'login')
    tenantPage = config.get('urlPath', 'tenant')
    createPage = config.get('urlPath', 'create')

    #print(keyring.get_credential(service_name='Onboard', username=None).username)
    #setup the chrome webdriver
    driver = setupDriver()
    driver.get(reactorLogin)

    #load previously saved cookies
    if (os.path.isfile('cookies.pkl')):
        cookies = pickle.load(open("cookies.pkl", "rb"))
        for cookie in cookies:
            driver.add_cookie(cookie)

    #login to the website
    login(driver, config)
    row_manual_review = []
    try:
        df = pd.read_excel(r'C:\Users\dmartinez\Documents\ope\testing.xlsx', sheet_name='Sheet1')
    except FileNotFoundError:
        print('bruh find yo file')
    except Exception as e:
        print(f"Error reading file: {e}")
        exit

    for index, row in df.iterrows():

        current_phone = str(row.get('PHONE', ''))
        current_email = str(row.get('EMAIL', ''))
        current_unit = str(row.get('UNIT', ''))
        current_first = str(row.get('FIRST', ''))
        current_last = str(row.get('LAST', ''))


        createEmail = False
        createNum = False
        manualEmail = False
        manualNum = False


        #Navigate to tenant search page
        driver.get(tenantPage)
        haltStep(2)

        #Locate and confirm search bar
        waitToLoad(driver, By.XPATH, config.get('searchPath', 'searchTenant'))
        search_element = driver.find_element(By.XPATH, config.get('searchPath', 'searchTenant'))
        haltStep(2)

        #fill input into tenant search bar and 'enter'
        search_element.send_keys(current_phone)
        haltStep(1)
        search_element.send_keys(Keys.ENTER)
        haltStep(1)

        #Check for the page reaction to search query
        try:
            #waitToLoad(driver,By.XPATH, config.get('searchPath', 'table'))
            driver.find_element(By.XPATH, config.get('searchPath', 'table'))
            haltStep(2)
            manualNum = True
        except NoSuchElementException:
            try:
                #waitToLoad(driver,By.XPATH, config.get('searchPath', 'notFound'))
                driver.find_element(By.XPATH, config.get('searchPath', 'notFound'))
                haltStep(2)
                createNum = True
            except NoSuchElementException:
                print("fkn skits mate")

        #fill input into tenant search bar and 'enter'
        search_element.send_keys(Keys.CONTROL + "a")
        haltStep(1)
        search_element.send_keys(Keys.BACK_SPACE)
        haltStep(1)
        search_element.send_keys(current_email)
        haltStep(1)
        search_element.send_keys(Keys.ENTER)
        haltStep(3)

        #Check for the page reaction to search query        
        try:
            #waitToLoad(driver,By.XPATH, config.get('searchPath', 'table'))
            table_element = driver.find_element(By.XPATH, config.get('searchPath', 'table'))
            haltStep(2)
            manualEmail = True
        except NoSuchElementException:
            try:
                #waitToLoad(driver,By.XPATH, config.get('searchPath', 'notFound'))
                notFound_element = driver.find_element(By.XPATH, config.get('searchPath', 'notFound'))
                haltStep(2)
                createEmail = True
            except NoSuchElementException:
                print("fkn skits mate")

        #waitToLoad(driver, By.XPATH, config.get('searchPath', 'searchTenant'))
        search_element.send_keys(Keys.CONTROL + "a")
        haltStep(1)
        search_element.send_keys(Keys.BACK_SPACE)
        haltStep(1)

        #create tenant
        if createNum == True and createEmail == True:

            #navigate to create tenant page
            driver.get(createPage);
            waitToLoad(driver, By.XPATH, config.get('createPath', 'SelectUnit'))
            haltStep(1)
            unit = driver.find_element(By.XPATH, config.get('createPath', 'SelectUnit'))
            unit.send_keys(Keys.CLEAR)
            unit.send_keys(current_unit)
            unit.send_keys(Keys.DOWN)
            unit.send_keys(Keys.ENTER)
            haltStep(1)
            driver.find_element(By.XPATH, config.get('createPath', 'first')).send_keys(current_first)
            driver.find_element(By.XPATH, config.get('createPath', 'last')).send_keys(current_last)
            driver.find_element(By.XPATH, config.get('createPath', 'address')).send_keys(current_email)
            driver.find_element(By.XPATH, config.get('createPath', 'number')).send_keys(current_phone)
            haltStep(2)
            driver.find_element(By.XPATH, config.get('createPath', 'createTenant')).click()
            haltStep(1)
            try:
                #waitToLoad(driver,By.XPATH, config.get('createPath', 'duplicate'))
                driver.find_element(By.XPATH, config.get('createPath', 'duplicate')).click()
                haltStep(1)
            except NoSuchElementException:
                print(current_first + current_last + "Created")
        
        elif manualEmail == True or manualNum == True:
            print("Move that mf")
            row_data = row.to_dict() # Get all columns from the current row
            row_data['ReasonForManualReview'] = []
            if manualNum:
                row_data['ReasonForManualReview'].append('Phone number already exists in system.')
            if manualEmail:
                row_data['ReasonForManualReview'].append('Email address already exists in system.')
            row_data['ReasonForManualReview'] = '; '.join(row_data['ReasonForManualReview']) # Join reasons
            row_manual_review.append(row_data) # Add the modified dictionary to the list
            print(f"Added row {index} to manual review list.")
            haltStep(1)
        else:
            print("move that mf")
            # This 'else' block catches cases where neither create conditions nor manual conditions are fully met.
            # It's good to capture these for debugging or further review.
            print(f"Case: Unhandled scenario for row {index}. Creating row for review.")
            row_data = row.to_dict()
            row_data['ReasonForManualReview'] = 'Unhandled scenario (neither creation nor clear manual case)'
            # You might want to distinguish this from explicit manualEmail/manualNum cases
            row_manual_review.append(row_data)
        haltStep(2)

            

        
    #dump the cookies back into the file
    pickle.dump(driver.get_cookies(), open('cookies.pkl', 'wb'))

    #driver quits
    driver.quit()

    writeOut(row_manual_review)

if __name__ == "__main__":
    main()